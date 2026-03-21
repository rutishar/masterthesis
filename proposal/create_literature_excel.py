"""Creates the literature screening Excel sheet for the supervisor."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

papers = [
    # ── RCPSP / Multi-Project Scheduling ─────────────────────────────────────
    {
        "authors": "Ratajczak-Ropel, E.; Jedrzejowicz, P.",
        "year": 2023,
        "title": "Parallelized Population-Based Multi-Heuristic System with Reinforcement Learning for Solving Multi-Skill RCPSP with Hierarchical Skills",
        "source": "FedCSIS (Conference on Computer Science and Information Systems)",
        "doi": "10.15439/2023f2826",
        "relevance": "High",
        "finding": "RL-controlled multi-heuristic system solves Multi-Skill RCPSP with hierarchical skill requirements — directly applicable to skill-aware CAPEX project staffing and allocation optimization.",
        "category": "RQ3 – Allocation Optimization",
    },
    {
        "authors": "Khajesaeedi, S.; Sadjadi, S.J.; Barzinpour, F.; Moghaddam, R.",
        "year": 2025,
        "title": "Resource-constrained project scheduling problem: Review of recent developments",
        "source": "Journal of Project Management",
        "doi": "10.5267/j.jpm.2024.12.002",
        "relevance": "High",
        "finding": "Comprehensive survey of RCPSP covering ML integration, multi-objective optimization, stochastic models, and hybrid metaheuristics — serves as key background reference for formulating CAPEX portfolio scheduling as an RCPSP variant.",
        "category": "RQ3 – Allocation Optimization",
    },
    {
        "authors": "van den Beek, T.; van Essen, J.T.; Pruyn, J.; Aardal, K.",
        "year": 2025,
        "title": "Machine learning assisted Differential Evolution for the Dynamic Resource Constrained Multi-project Scheduling Problem with Static project Schedules",
        "source": "European Journal of Operational Research",
        "doi": "10.1016/j.ejor.2025.05.059",
        "relevance": "High",
        "finding": "ML-guided differential evolution solves dynamic multi-project RCPSP — demonstrates that ML-assisted metaheuristics outperform classical approaches for portfolio-level resource scheduling, directly supporting the thesis approach.",
        "category": "RQ3 – Allocation Optimization",
    },
    {
        "authors": "Liu, H.; Zhang, J.; Demeulemeester, E.; Chen, Z.",
        "year": 2025,
        "title": "A machine learning-based genetic programming algorithm for the stochastic RCPSP with 3D workspaces and carbon emissions",
        "source": "International Journal of Production Research",
        "doi": "10.1080/00207543.2025.2583475",
        "relevance": "Medium",
        "finding": "ML-based genetic programming for stochastic RCPSP — uncertainty modelling methodology is applicable to CAPEX demand forecasting under project variability; sustainability objectives are relevant for infrastructure CAPEX.",
        "category": "RQ3 – Allocation Optimization",
    },
    {
        "authors": "Dong, Y.; Tang, L.; Jia, W.",
        "year": 2025,
        "title": "A Learning-assisted Discrete Differential Evolution for Resource Constrained Project Scheduling",
        "source": "GECCO – Annual Conference on Genetic and Evolutionary Computation",
        "doi": "10.1145/3712256.3726397",
        "relevance": "Medium",
        "finding": "k-NN surrogate model accelerates evolutionary search for RCPSP — methodological contribution applicable to skill-aware CAPEX allocation heuristics.",
        "category": "RQ3 – Allocation Optimization",
    },
    # ── Knowledge Graphs & Workforce ─────────────────────────────────────────
    {
        "authors": "Fettach, Y.; Bahaj, A.B.; Ghogho, M.",
        "year": 2025,
        "title": "Skill Demand Forecasting Using Temporal Knowledge Graph Embeddings",
        "source": "arXiv",
        "doi": "10.48550/arXiv.2504.07233",
        "relevance": "High",
        "finding": "Temporal KG embeddings from job advertisements forecast skill demand via link prediction — the most directly relevant paper to the thesis: temporal KG + forecasting for skill demand is the core thesis concept applied to an adjacent domain.",
        "category": "RQ1 – Demand Forecasting / RQ2 – Knowledge Representation",
    },
    {
        "authors": "Agrawal, S.; Vats, R.",
        "year": 2025,
        "title": "Dynamic Knowledge Graphs: Revolutionizing Skill Analytics through Graph Neural Networks",
        "source": "International Journal of Scientific Research in Computer Science, Engineering and Information Technology",
        "doi": "10.32628/cseit251112181",
        "relevance": "High",
        "finding": "GNN-based dynamic KG for hierarchical skill modelling with temporal trend analysis — directly applicable to the knowledge representation layer of the CAPEX resource planning system.",
        "category": "RQ2 – Knowledge Representation",
    },
    {
        "authors": "Seif, A.; Toh, S.; Lee, H.K.",
        "year": 2024,
        "title": "A Dynamic Jobs-Skills Knowledge Graph",
        "source": "HR@RecSys Workshop",
        "doi": "",
        "relevance": "Medium",
        "finding": "Dynamic modelling of job-skill relationships via KGs for recommendation — applicable to matching engineers to CAPEX project roles in real time.",
        "category": "RQ2 – Knowledge Representation",
    },
    {
        "authors": "Tosic, M.; Petrovic, N.; Tosic, O.",
        "year": 2024,
        "title": "Workforce wellbeing management leveraging semantic knowledge graph",
        "source": "Zbornik radova (Conference Proceedings)",
        "doi": "10.5937/imcsm24042t",
        "relevance": "Low",
        "finding": "Semantic KGs for fatigue-aware workforce scheduling integrated with ERP — shows KG-ERP integration patterns applicable to the Run phase of the thesis system.",
        "category": "RQ2 – Knowledge Representation",
    },
    # ── ESCO Ontology ─────────────────────────────────────────────────────────
    {
        "authors": "Sun, Y.",
        "year": 2026,
        "title": "Contrastive Bi-Encoder Models for Multi-Label Skill Extraction: Enhancing ESCO Ontology Matching with BERT and Attention Mechanisms",
        "source": "arXiv",
        "doi": "10.48550/arXiv.2601.09119",
        "relevance": "High",
        "finding": "Zero-shot ESCO skill extraction from job descriptions using contrastive bi-encoders (F1@5 = 0.72) — directly enables automated skill tagging of CAPEX project descriptions without manual annotation.",
        "category": "RQ2 – Knowledge Representation",
    },
    {
        "authors": "Criscuolo, S. et al.",
        "year": 2025,
        "title": "Towards Machine Learning-Based Ontology Mapping to Bridge O*NET and ESCO Skills",
        "source": "IEEE MetroXRAINE 2025",
        "doi": "10.1109/MetroXRAINE66377.2025.11340152",
        "relevance": "High",
        "finding": "96.8% precision ontology alignment between ESCO and O*NET via word embeddings — important for a CAPEX resource platform integrating skill data from heterogeneous HR systems using different standards.",
        "category": "RQ2 – Knowledge Representation",
    },
    {
        "authors": "Saraswati, F.; Baizal, Z.",
        "year": 2025,
        "title": "Hierarchical TF-IDF for Tech Job Recommendation: Comparing ESCO and MIND Ontologies",
        "source": "BTS-I2C 2025 Conference",
        "doi": "10.1109/BTS-I2C67944.2025.11399358",
        "relevance": "Medium",
        "finding": "Domain-specific ontology achieves 418% precision improvement over generic ESCO for skill matching — suggests a CAPEX-domain extension of ESCO may be necessary for high-precision resource recommendations.",
        "category": "RQ2 – Knowledge Representation",
    },
    {
        "authors": "Zhang, Y.",
        "year": 2025,
        "title": "Semantic-web-Enhanced Hybrid Learning for Career Planning: Ontology-driven Matching, Sequence Forecasting, and Closed-loop Optimization",
        "source": "Journal of ICT Standardization",
        "doi": "10.13052/jicts2245-800x.1334",
        "relevance": "High",
        "finding": "Combines domain ontology + gradient-boosted trees + transformer-based sequence models for skill forecasting and matching — hybrid AI architecture directly parallels the thesis approach.",
        "category": "RQ1 – Demand Forecasting / RQ2 – Knowledge Representation",
    },
    # ── CAPEX / Portfolio / Capital Allocation ────────────────────────────────
    {
        "authors": "Anonymous",
        "year": 2023,
        "title": "Conditional Portfolio Optimization: Using Machine Learning to Adapt Capital Allocations to Market Regimes",
        "source": "SSRN Working Paper",
        "doi": "10.2139/ssrn.4383184",
        "relevance": "Medium",
        "finding": "ML-driven adaptive capital allocation conditioned on regime detection — methodology applicable to dynamic CAPEX budget reallocation across portfolio phases.",
        "category": "RQ4 – Evaluation / Baseline",
    },
    {
        "authors": "Huang, S.",
        "year": 2024,
        "title": "Advancing portfolio optimization: The convergence of machine learning and traditional financial models",
        "source": "Applied and Computational Engineering",
        "doi": "10.54254/2755-2721/57/20241335",
        "relevance": "Medium",
        "finding": "RL applied to classical portfolio models (CAPM/APT) for dynamic allocation — RL-based allocation methodology transferable to CAPEX resource allocation under uncertainty.",
        "category": "RQ3 – Allocation Optimization",
    },
    # ── Foundational References ───────────────────────────────────────────────
    {
        "authors": "Hevner, A.R.; March, S.T.; Park, J.; Ram, S.",
        "year": 2004,
        "title": "Design Science in Information Systems Research",
        "source": "MIS Quarterly",
        "doi": "10.2307/25148625",
        "relevance": "High",
        "finding": "Establishes DSR as the rigorous methodology for building and evaluating IT artifacts — provides the epistemological foundation for the thesis research paradigm.",
        "category": "Research Design",
    },
    {
        "authors": "Peter, M.K.; Laurenzi, E.; Hinkelmann, K.",
        "year": 2025,
        "title": "Workshop Canvas AI-4-SME Framework",
        "source": "FHNW School of Business",
        "doi": "",
        "relevance": "High",
        "finding": "Three-phase AI framework (Design/Build/Run) with KIA/DIA distinction and hybrid AI process model — adopted as the overarching methodological structure of this thesis.",
        "category": "Research Design / RQ2",
    },
    {
        "authors": "Garcez, A.; Lamb, L.C.",
        "year": 2022,
        "title": "Neural-Symbolic Learning and Reasoning: A Survey and Interpretation",
        "source": "Neuro-Symbolic Artificial Intelligence: The State of the Art (IOS Press)",
        "doi": "",
        "relevance": "High",
        "finding": "Comprehensive survey of neuro-symbolic AI combining symbolic reasoning with neural learning — theoretical foundation for the hybrid AI approach combining ESCO knowledge graphs with ML forecasting models.",
        "category": "RQ2 – Knowledge Representation",
    },
    {
        "authors": "Lim, B.; Arik, S.O.; Loeff, N.; Pfister, T.",
        "year": 2021,
        "title": "Temporal Fusion Transformers for Interpretable Multi-horizon Time Series Forecasting",
        "source": "International Journal of Forecasting",
        "doi": "10.1016/j.ijforecast.2021.03.012",
        "relevance": "High",
        "finding": "TFT combines multi-head attention with interpretable variable selection for multi-horizon forecasting — strong candidate architecture for temporal skill-demand forecasting across CAPEX project phases.",
        "category": "RQ1 – Demand Forecasting",
    },
    {
        "authors": "Chen, T.; Guestrin, C.",
        "year": 2016,
        "title": "XGBoost: A Scalable Tree Boosting System",
        "source": "Proceedings of the 22nd ACM SIGKDD",
        "doi": "10.1145/2939672.2939785",
        "relevance": "High",
        "finding": "XGBoost achieves state-of-the-art performance on tabular data with gradient boosting — baseline model for skill-demand forecasting from structured project features.",
        "category": "RQ1 – Demand Forecasting",
    },
    {
        "authors": "Demeulemeester, E.L.; Herroelen, W.S.",
        "year": 2002,
        "title": "Project Scheduling: A Research Handbook",
        "source": "Springer",
        "doi": "10.1007/978-1-4615-1003-4",
        "relevance": "High",
        "finding": "Definitive reference on project scheduling theory including RCPSP formulations — provides the formal background for the resource allocation optimization model in the thesis.",
        "category": "RQ3 – Allocation Optimization",
    },
]

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Literature Screening"

# Colors
header_fill   = PatternFill("solid", fgColor="1F4E79")
high_fill     = PatternFill("solid", fgColor="C6EFCE")
medium_fill   = PatternFill("solid", fgColor="FFEB9C")
low_fill      = PatternFill("solid", fgColor="FFC7CE")
alt_fill      = PatternFill("solid", fgColor="F2F2F2")

header_font   = Font(bold=True, color="FFFFFF", size=11)
body_font     = Font(size=10)
wrap_align    = Alignment(wrap_text=True, vertical="top")
center_align  = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

headers = ["#", "Authors", "Year", "Title", "Source / Journal", "DOI",
           "Relevance", "Finding", "Category", "Notes"]
col_widths = [4, 32, 6, 55, 30, 38, 11, 60, 30, 20]

# Header row
for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

# Data rows
for i, p in enumerate(papers, start=1):
    row = i + 1
    values = [i, p["authors"], p["year"], p["title"], p["source"],
              p["doi"], p["relevance"], p["finding"], p["category"], ""]

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = center_align if col_idx in (1, 3, 7) else wrap_align

    # Relevance color
    rel_cell = ws.cell(row=row, column=7)
    if p["relevance"] == "High":
        rel_cell.fill = high_fill
    elif p["relevance"] == "Medium":
        rel_cell.fill = medium_fill
    else:
        rel_cell.fill = low_fill

    # Alternating row background (only non-relevance cols)
    if i % 2 == 0:
        for col_idx in [1, 2, 3, 4, 5, 6, 8, 9, 10]:
            ws.cell(row=row, column=col_idx).fill = alt_fill

    ws.row_dimensions[row].height = 80

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "Literature Screening Summary"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")

totals = {"High": 0, "Medium": 0, "Low": 0}
for p in papers:
    totals[p["relevance"]] += 1

ws2["A3"] = "Relevance"
ws2["B3"] = "Count"
ws2["A3"].font = Font(bold=True)
ws2["B3"].font = Font(bold=True)
ws2["A4"] = "High"
ws2["B4"] = totals["High"]
ws2["A4"].fill = high_fill
ws2["A5"] = "Medium"
ws2["B5"] = totals["Medium"]
ws2["A5"].fill = medium_fill
ws2["A6"] = "Low"
ws2["B6"] = totals["Low"]
ws2["A6"].fill = low_fill
ws2["A7"] = "Total"
ws2["B7"] = sum(totals.values())
ws2["A7"].font = Font(bold=True)
ws2["B7"].font = Font(bold=True)

for col in ["A", "B"]:
    ws2.column_dimensions[col].width = 18

out_path = r"C:\Users\ramon\00_no_sync\masterthesis\proposal\literature_screening.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Papers: {sum(totals.values())} total — {totals['High']} High, {totals['Medium']} Medium, {totals['Low']} Low")
