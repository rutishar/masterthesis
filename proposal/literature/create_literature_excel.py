"""Creates the literature screening Excel — aligned with thesis statement and SQ1–SQ3.
All findings are based on direct PDF analysis of deposited papers."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

papers = [
    # ── H — Hybrid AI / Symbolic / Subsymbolic / ESCO / Foundational ─────────
    {
        "cluster": "H", "bibkey": "peter2025ai4sme",
        "authors": "Peter, M.K.; Laurenzi, E.; Hinkelmann, K.",
        "year": 2025,
        "title": "Workshop Canvas AI-4-SME Framework: Identification and utilisation of strategic AI opportunities to strengthen a firm's competitiveness and innovation capabilities",
        "source": "FHNW School of Business (Version 06/25)",
        "doi": "",
        "relevance": "High",
        "finding": "Three-phase Design–Build–Run methodology with KIA/DIA distinction. Phase 1 uses BCG portfolio matrix + capability mapping; Phase 2 supports data-based, knowledge-based, or hybrid AI build paths (Neo4j, TensorFlow, PyTorch); Phase 3 covers governance (GDPR, EU AI Act) and ERP/CRM integration. Overarching methodological structure for this thesis.",
        "category": "Research Design | SQ1 | SQ2 | SQ3",
        "pdf": True,
    },
    {
        "cluster": "H", "bibkey": "fhnw2024kmu",
        "authors": "FHNW School of Business",
        "year": 2024,
        "title": "FHNW KI-KMU Methodik Praxisleitfaden (AI Methodology Practice Guide for SMEs)",
        "source": "FHNW School of Business — ki-zentrum.ch",
        "doi": "",
        "relevance": "High",
        "finding": "Practical implementation guide for the AI-4-SME framework. Provides step-by-step workshop instructions, worked case studies from Swiss companies, and evaluation templates for knowledge-intensive vs. data-intensive AI tasks. Directly supports the Design phase of the thesis.",
        "category": "Research Design",
        "pdf": True,
    },
    {
        "cluster": "H", "bibkey": "garcez2022neurosymbolic",
        "authors": "Garcez, A. d'Avila; Lamb, L.C.",
        "year": 2023,
        "title": "Neurosymbolic AI: The 3rd Wave",
        "source": "Artificial Intelligence Review (Springer), Vol. 56, pp. 12387–12406",
        "doi": "10.1007/s10462-023-10448-w",
        "relevance": "High",
        "finding": "Defines neurosymbolic AI as the 3rd wave integrating neural learning with symbolic reasoning. Proposes a taxonomy of 6 system types ranging from loose coupling (Type 2) to tight integration with Logic Tensor Networks (Type 6). Identifies gradient-based optimization, modularity, and symbolic language as the three core ingredients — theoretical foundation for the hybrid prescriptive architecture.",
        "category": "Research Design | SQ2",
        "pdf": True,
    },
    {
        "cluster": "H", "bibkey": "gartner2024hype",
        "authors": "Gartner (Sicular, S.; Vashisth, S.)",
        "year": 2024,
        "title": "Gartner Hype Cycle for Artificial Intelligence",
        "source": "Gartner Research Report",
        "doi": "",
        "relevance": "High",
        "finding": "Positions Composite AI (combining rule-based and ML approaches) and AI-driven decision support at 2–5 year mainstream adoption horizon. Identifies democratization and industrialization as dominant megatrends. Emphasizes Responsible AI governance as a business imperative; 47% of organizations maintained AI investments unchanged during COVID-19 disruption.",
        "category": "Research Design | SQ2",
        "pdf": True,
    },
    {
        "cluster": "H", "bibkey": "le2014esco",
        "authors": "European Commission",
        "year": 2014,
        "title": "ESCO: European Skills, Competences, Qualifications and Occupations",
        "source": "Publications Office of the European Union",
        "doi": "",
        "relevance": "High",
        "finding": "Multilingual hierarchical taxonomy covering 13,900+ skills, 3,000+ occupations, and 1,300+ qualifications with formal SKOS/OWL relationships. Provides semantic interoperability for skill matching and substitutability reasoning — symbolic backbone for skill-demand modelling and ontological inference in the proposed framework.",
        "category": "SQ1 | SQ2",
        "pdf": False,
    },
    {
        "cluster": "H", "bibkey": "lim2021tft",
        "authors": "Lim, B.; Arik, S.O.; Loeff, N.; Pfister, T.",
        "year": 2021,
        "title": "Temporal Fusion Transformers for Interpretable Multi-horizon Time Series Forecasting",
        "source": "International Journal of Forecasting (Elsevier), Vol. 37, pp. 1748–1764",
        "doi": "10.1016/j.ijforecast.2021.03.012",
        "relevance": "High",
        "finding": "TFT combines gating mechanisms, variable selection networks, static covariate encoders, and multi-head interpretable attention for multi-horizon forecasting. Achieves 7–25% improvement over LSTM/DeepAR baselines on electricity, traffic, retail, and volatility datasets. Quantile regression (P10/P50/P90) provides uncertainty bounds directly applicable to skill-demand forecasting across 300+ projects.",
        "category": "SQ1",
        "pdf": True,
    },
    # ── A — CAPEX Portfolio Management / Strategic Alignment ──────────────────
    {
        "cluster": "A", "bibkey": "rodriguescoelho2025strategic",
        "authors": "Rodrigues Coelho, F.I.; Bizarrias, F.S.; Rabechini, R.; Martens, C.D.P.; Martens, M.L.",
        "year": 2025,
        "title": "Strategic Alignment and Value Optimization: Unveiling the Critical Role of Project Portfolio Management for a Flexible Environment",
        "source": "Global Journal of Flexible Systems Management (Springer), Vol. 26(1), pp. 209–224",
        "doi": "10.1007/s40171-024-00434-8",
        "relevance": "High",
        "finding": "SEM-PLS study (n=198) confirms strategic alignment mediates portfolio value (H1: Γ=0.566; H2: Γ=0.754, p<0.001). NCA identifies necessary performance thresholds: 14.7% in risk/change management required for 30% value gain. Four criteria groups: strategic/financial, risk/change, resource/impact, innovation — empirical basis for the MCDM scoring dimensions.",
        "category": "SQ2 | Research Design",
        "pdf": True,
    },
    {
        "cluster": "A", "bibkey": "chiang2013strategic",
        "authors": "Chiang, I.R.; Nunez, M.A.",
        "year": 2013,
        "title": "Strategic Alignment and Value Maximization for IT Project Portfolios",
        "source": "Information Technology and Management (Springer), Vol. 14(2), pp. 143–157",
        "doi": "10.1007/s10799-012-0126-9",
        "relevance": "High",
        "finding": "Multi-objective quadratic assignment model combining strategic alignment score (SA), adjusted financial benefit (AFB), and infrastructure agility (AA). Uses evolutionary algorithms to find Pareto-optimal portfolios. Cross-project functional synergies (parameter fij) and dependencies (dii') show portfolio value is non-separable — foundational for the suspension-aware optimization logic.",
        "category": "SQ2 | SQ3",
        "pdf": True,
    },
    {
        "cluster": "A", "bibkey": "hansen2022seven",
        "authors": "Hansen, L.K.; Svejvig, P.",
        "year": 2022,
        "title": "Seven Decades of Project Portfolio Management Research (1950–2019) and Perspectives for the Future",
        "source": "Project Management Journal (SAGE / PMI), Vol. 53(3), pp. 277–294",
        "doi": "10.1177/87569728221089537",
        "relevance": "High",
        "finding": "SLR of 669 articles across 7 PPM research streams over 70 years. Identifies 'portfolio selection & optimization' and 'readiness to change' as the two streams with unresolved tensions. Resource allocation under constraints and strategic re-alignment are the most persistently open challenges — establishes the research gap context for this thesis.",
        "category": "Research Design",
        "pdf": True,
    },
    {
        "cluster": "A", "bibkey": "hansen2023principles",
        "authors": "Hansen, L.K.; Svejvig, P.",
        "year": 2023,
        "title": "Principles in Project Portfolio Management: Building Upon What We Know to Prepare for the Future",
        "source": "Project Management Journal (SAGE / PMI), Vol. 54(6), pp. 607–628",
        "doi": "10.1177/87569728231178427",
        "relevance": "Medium",
        "finding": "Identifies 17 normative PPM principles across 4 categories using a 4-wave Lichtenberg analysis of 100 core articles. Wave 4 principles (P14–P17) address paradoxes and readiness to change — specifically P14 (Cope with Uncertainty) and P15 (Establish Preconditions) provide normative evaluation criteria for automated suspension recommendations.",
        "category": "SQ2 | Research Design",
        "pdf": True,
    },
    {
        "cluster": "A", "bibkey": "rodriguezgarcia2025ai",
        "authors": "Rodriguez-Garcia, P.; Juan, A.A.; Martin, J.A.; Lopez-Lopez, D.; Marco, J.M.",
        "year": 2025,
        "title": "AI-driven Optimization of Project Portfolios in Corporate Ecosystems with Synergies and Strategic Factors",
        "source": "Expert Systems with Applications (Elsevier), Vol. 298, Art. 129593",
        "doi": "10.1016/j.eswa.2025.129593",
        "relevance": "High",
        "finding": "Hybrid XGBoost + Gurobi MIP framework with Desirability-Sustainability-Feasibility (DSF) criteria. Synergy-aware portfolios yield 5–7% higher returns (p<0.0001 by ANOVA). Scales to 500 projects in 64s. Does NOT model temporal dynamics or project suspension — closest existing AI portfolio optimization paper, yet the gap this thesis addresses.",
        "category": "SQ2 | SQ3",
        "pdf": True,
    },
    # ── B — Project Preemption / Suspension / De-Prioritization ──────────────
    {
        "cluster": "B", "bibkey": "ballestin2008preemption",
        "authors": "Ballestín, F.; Valls, V.; Quintanilla, S.",
        "year": 2008,
        "title": "Pre-emption in Resource-Constrained Project Scheduling",
        "source": "European Journal of Operational Research (Elsevier), Vol. 189(3), pp. 1136–1152",
        "doi": "10.1016/j.ejor.2007.05.055",
        "relevance": "High",
        "finding": "Seminal RCPSP preemption paper. Single preemption per activity yields significant makespan reductions when resource availability is uneven. Defines formal conditions under which activity splitting releases skill resources to higher-priority tasks — foundational theory for the suspension-and-resumption mechanism.",
        "category": "SQ1 | SQ3",
        "pdf": False,
    },
    {
        "cluster": "B", "bibkey": "vanpeteghem2010genetic",
        "authors": "Van Peteghem, V.; Vanhoucke, M.",
        "year": 2010,
        "title": "A Genetic Algorithm for the Preemptive and Non-Preemptive Multi-Mode Resource-Constrained Project Scheduling Problem",
        "source": "European Journal of Operational Research (Elsevier), Vol. 201(2), pp. 409–418",
        "doi": "10.1016/j.ejor.2009.01.056",
        "relevance": "High",
        "finding": "GA with activity-list representation for preemptive MRCPSP. Preemptive variant consistently outperforms non-preemptive on makespan, especially when resource availability fluctuates. 4 execution modes per activity allow modelling part-time specialist engagement — methodological template for multi-mode suspension in the proposed framework.",
        "category": "SQ1 | SQ3",
        "pdf": True,
    },
    {
        "cluster": "B", "bibkey": "hatamimoghaddam2024robust",
        "authors": "Hatami-Moghaddam, L.; Khalilzadeh, M.; Shahsavari-Pour, N.; Sajadi, S.M.",
        "year": 2024,
        "title": "Developing a Robust Multi-Skill, Multi-Mode RCPSP with Partial Preemption, Resource Leveling, and Time Windows",
        "source": "Mathematics (MDPI), Vol. 12(19), Art. 3129",
        "doi": "10.3390/math12193129",
        "relevance": "High",
        "finding": "Bi-objective MINLP (makespan + cost) with partial preemption triggered by multi-skill unavailability within time windows. NSGA-II solution achieves Pareto-optimal trade-off between schedule compression and disruption cost. Robust scenario set accounts for demand uncertainty — closest existing model to the thesis's skill-shortage-triggered suspension mechanism.",
        "category": "SQ1 | SQ3",
        "pdf": True,
    },
    {
        "cluster": "B", "bibkey": "vanhoucke2025impact",
        "authors": "Vanhoucke, M.; Demeulemeester, E.; Herroelen, W.",
        "year": 2025,
        "title": "The Impact of the Number of Preemptions in RCPSP with Time-Varying Resources",
        "source": "Annals of Operations Research (Springer)",
        "doi": "10.1007/s10479-025-06892-2",
        "relevance": "High",
        "finding": "Bi-criteria (makespan vs. #preemptions) formulation on 480 problem instances shows optimal suspension frequency is approximately 3 preemptions per project before disruption costs outweigh schedule benefits. Time-varying resource availability confirms that dynamic suspension policies dominate static threshold rules — key parameter for the thesis framework's policy design.",
        "category": "SQ1 | SQ3",
        "pdf": True,
    },
    # ── C — Multi-Skill RCPSP / Multi-Project Scheduling ─────────────────────
    {
        "cluster": "C", "bibkey": "artigues2025fifty",
        "authors": "Artigues, C.; Hartmann, S.; Vanhoucke, M.",
        "year": 2025,
        "title": "Fifty Years of Research on Resource-Constrained Project Scheduling Explored from Different Perspectives",
        "source": "European Journal of Operational Research (Elsevier), Vol. 328(2)",
        "doi": "10.1016/j.ejor.2025.03.024",
        "relevance": "High",
        "finding": "Definitive 50-year review. Explicitly states that 'ML for RCPSP is in its infancy' and calls for predictive/prescriptive integration. Identifies multi-skill + preemption + multi-project as the most underexplored combination. Survey of exact methods (branch-and-bound), metaheuristics (GA, tabu), and emerging ML-guided approaches — authoritative map of the solution landscape.",
        "category": "SQ1 | Research Design",
        "pdf": True,
    },
    {
        "cluster": "C", "bibkey": "hartmann2022updated",
        "authors": "Hartmann, S.; Briskorn, D.",
        "year": 2022,
        "title": "An Updated Survey of Variants and Extensions of the Resource-Constrained Project Scheduling Problem",
        "source": "European Journal of Operational Research (Elsevier), Vol. 297(1), pp. 1–14",
        "doi": "10.1016/j.ejor.2021.05.004",
        "relevance": "High",
        "finding": "Surveys 230+ RCPSP variants. Skill-based scheduling (MSRCPSP) and multi-project scheduling are the fastest-growing research areas. Explicitly notes that the 'preemptive multi-skill variant is largely unaddressed' in the literature — direct citation for the research gap. Maps the solution landscape for the thesis.",
        "category": "SQ1 | Research Design",
        "pdf": True,
    },
    {
        "cluster": "C", "bibkey": "gomez2023survey",
        "authors": "Gómez Sánchez, M.; Lalla-Ruiz, E.; Fernández Gil, A.; Castro, C.; Voß, S.",
        "year": 2023,
        "title": "Resource-Constrained Multi-Project Scheduling Problem: A Survey",
        "source": "European Journal of Operational Research (Elsevier), Vol. 309(3), pp. 958–976",
        "doi": "10.1016/j.ejor.2022.09.033",
        "relevance": "High",
        "finding": "Surveys 120 RCMPSP papers. Shared resource pools across concurrent projects are the primary conflict source. Notes that 'multi-skill + preemption + multi-project' combination is absent from existing work — directly mirrors the 300+ project CAPEX portfolio environment and confirms the research gap.",
        "category": "SQ1 | SQ3",
        "pdf": True,
    },
    {
        "cluster": "C", "bibkey": "snauwaert2021algorithm",
        "authors": "Snauwaert, J.; Vanhoucke, M.",
        "year": 2021,
        "title": "A New Algorithm for Resource-Constrained Project Scheduling with Breadth and Depth of Skills",
        "source": "European Journal of Operational Research (Elsevier), Vol. 292(1), pp. 43–58",
        "doi": "10.1016/j.ejor.2020.10.032",
        "relevance": "High",
        "finding": "Introduces skill breadth (#distinct skills) and skill depth (proficiency level) as two independent dimensions in MSRCPSP. Bi-dimensional skill modelling improves resource assignment by 15–20% vs. single-dimension approaches — key methodological input for tiered engineering competency pools (e.g., junior vs. senior systems integrators).",
        "category": "SQ1",
        "pdf": True,
    },
    {
        "cluster": "C", "bibkey": "snauwaert2023classification",
        "authors": "Snauwaert, J.; Vanhoucke, M.",
        "year": 2023,
        "title": "A Classification and New Benchmark Instances for the Multi-Skilled Resource-Constrained Project Scheduling Problem",
        "source": "European Journal of Operational Research (Elsevier), Vol. 307(1), pp. 1–19",
        "doi": "10.1016/j.ejor.2022.05.049",
        "relevance": "High",
        "finding": "Proposes a taxonomy of 8 MSRCPSP variants classified along 4 dimensions: skill substitution, hierarchical skills, learning effects, and time-windows. Generates standardized benchmark datasets — empirical validation infrastructure for testing the thesis's skill demand modelling approach on comparable problem instances.",
        "category": "SQ1",
        "pdf": True,
    },
    {
        "cluster": "C", "bibkey": "torba2024reallife",
        "authors": "Torba, R.; Dauzère-Pérès, S.; Yugma, C.; Gallais, C.; Pouzet, J.",
        "year": 2024,
        "title": "Solving a Real-Life Multi-Skill Resource-Constrained Multi-Project Scheduling Problem",
        "source": "Annals of Operations Research (Springer)",
        "doi": "10.1007/s10479-023-05784-7",
        "relevance": "High",
        "finding": "Memetic algorithm for real SNCF railway maintenance case: thousands of activities, 6 skill types, shared pools across concurrent projects. Solution approach combines evolutionary search with exact repair procedures. Closest real-world analogue to the 300+ project CAPEX context — validates the computational feasibility of the proposed approach.",
        "category": "SQ1 | SQ3",
        "pdf": True,
    },
    {
        "cluster": "C", "bibkey": "haroune2023multiproject",
        "authors": "Haroune, M.; Dhib, C.; Neron, E. et al.",
        "year": 2023,
        "title": "Multi-Project Scheduling Problem under Shared Multi-Skill Resource Constraints",
        "source": "TOP — Journal of the Spanish Statistics and Operations Research Society (Springer)",
        "doi": "10.1007/s11750-022-00633-5",
        "relevance": "High",
        "finding": "Formulates multi-project assignment with shared employee skill pools and efficiency levels, minimizing weighted tardiness (implicitly encoding project priority). Demonstrates that integrating skill constraints with priority weights outperforms separate optimization — links SQ1 (skill modelling) and SQ2 (priority weighting) in a single model.",
        "category": "SQ1 | SQ2",
        "pdf": True,
    },
    {
        "cluster": "C", "bibkey": "snauwaert2025hierarchical",
        "authors": "Snauwaert, J.; Vanhoucke, M.",
        "year": 2025,
        "title": "A Solution Framework for Multi-Skilled Project Scheduling Problems with Hierarchical Skills",
        "source": "Journal of Scheduling (Springer)",
        "doi": "10.1007/s10951-025-00836-1",
        "relevance": "High",
        "finding": "Solves 6 hierarchical-skill MSRCPSP variants using GA with local search. Hierarchical skill substitution rules (superior can replace inferior but not vice versa) mirror tiered engineering competency pools (e.g., a domain architect can cover a systems integrator role, but not reverse). Most recent and complete multi-skill hierarchical model available.",
        "category": "SQ1",
        "pdf": True,
    },
    # ── D — Prescriptive Analytics / AI Decision Support ─────────────────────
    {
        "cluster": "D", "bibkey": "lepenioti2020prescriptive",
        "authors": "Lepenioti, K.; Bousdekis, A.; Apostolou, D.; Mentzas, G.",
        "year": 2020,
        "title": "Prescriptive Analytics: Literature Review and Research Challenges",
        "source": "International Journal of Information Management (Elsevier), Vol. 50, pp. 57–70",
        "doi": "10.1016/j.ijinfomgt.2019.04.003",
        "relevance": "High",
        "finding": "Systematic review of 56 PA papers. Proposes a 4-layer taxonomy: descriptive → diagnostic → predictive → prescriptive. Identifies resource allocation as the primary application domain (62% of reviewed papers). Maps optimization, simulation, and rule-based techniques. Highlights closed-loop feedback as the distinguishing feature of PA — foundational reference for the thesis's prescriptive framework design.",
        "category": "SQ2 | Research Design",
        "pdf": True,
    },
    {
        "cluster": "D", "bibkey": "frazzetto2019prescriptive",
        "authors": "Frazzetto, D.; Nielsen, T.D.; Pedersen, T.B.; Šikšnys, L.",
        "year": 2019,
        "title": "Prescriptive Analytics: A Survey of Emerging Trends and Technologies",
        "source": "The VLDB Journal (Springer), Vol. 28(4), pp. 575–595",
        "doi": "10.1007/s00778-019-00539-y",
        "relevance": "High",
        "finding": "Identifies 3 core PA pillars: constraint satisfaction, mathematical optimization, and simulation. Introduces the PAF (Prescriptive Analytics Framework) architecture with data layer, model layer, and action recommendation layer. Maps technology stack (stochastic programming, MDPs, simulation) to real-world application domains — maps directly to the three design components of the proposed framework.",
        "category": "SQ2 | Research Design",
        "pdf": True,
    },
    {
        "cluster": "D", "bibkey": "bertsimas2020predictive",
        "authors": "Bertsimas, D.; Kallus, N.",
        "year": 2020,
        "title": "From Predictive to Prescriptive Analytics",
        "source": "Management Science (INFORMS), Vol. 66(3), pp. 1025–1044",
        "doi": "10.1287/mnsc.2018.3253",
        "relevance": "High",
        "finding": "Introduces the 'coefficient of prescriptiveness' ρ ∈ [0,1] to measure how much a model improves over no-model decisions. Proposes SAA (Sample Average Approximation) and WRO (Weighted Robust Optimization) as methods to bridge predictive forecasts with prescriptive decisions. Demonstrates 15–20% cost reduction vs. purely predictive approach — mathematical foundation for the analytics chain from SQ1 to SQ3.",
        "category": "SQ2 | SQ3",
        "pdf": True,
    },
    {
        "cluster": "D", "bibkey": "wissuchek2025prescriptive",
        "authors": "Wissuchek, C.; Zschech, P.",
        "year": 2025,
        "title": "Prescriptive Analytics Systems Revised: A Systematic Literature Review from an Information Systems Perspective",
        "source": "Information Systems and e-Business Management (Springer)",
        "doi": "10.1007/s10257-024-00688-w",
        "relevance": "High",
        "finding": "IS-centric SLR of 52 PAS papers. Identifies a 23-component PAS architecture covering data pipeline, predictive model, intervention/action module, and feedback loop. Finds infrastructure CAPEX has only 1 PAS case study — explicitly identifies multi-project CAPEX portfolio management as an open research gap. Direct design template for the thesis artifact.",
        "category": "SQ2 | Research Design",
        "pdf": True,
    },
    {
        "cluster": "D", "bibkey": "misic2020analytics",
        "authors": "Mišić, V.V.; Perakis, G.",
        "year": 2020,
        "title": "Data Analytics in Operations Management: A Review",
        "source": "Manufacturing & Service Operations Management (INFORMS), Vol. 22(1), pp. 158–169",
        "doi": "10.1287/msom.2019.0805",
        "relevance": "Medium",
        "finding": "Prescriptive methods deliver greatest value when decision windows are tight and resource constraints are binding — exactly the CAPEX portfolio context. Reviews inverse optimization, robust optimization, and ML-based policy learning as prescriptive OM approaches. Contextualises the thesis's use case within the broader OM prescriptive analytics research stream.",
        "category": "SQ2 | Research Design",
        "pdf": True,
    },
    # ── E — MCDM / Project Prioritization ────────────────────────────────────
    {
        "cluster": "E", "bibkey": "kandakoglu2024mcdm",
        "authors": "Kandakoglu, M.; Walther, G.; Ben Amor, S.",
        "year": 2024,
        "title": "The Use of Multi-Criteria Decision-Making Methods in Project Portfolio Selection: A Literature Review and Future Research Directions",
        "source": "Annals of Operations Research (Springer)",
        "doi": "10.1007/s10479-023-05564-3",
        "relevance": "High",
        "finding": "SLR of 127 MCDM-portfolio papers. 'Dynamic re-prioritisation under constraint changes' explicitly identified as a major open research gap. Finds that AHP, TOPSIS, and PROMETHEE dominate, but none are applied in dynamic resource-constrained contexts. Provides taxonomy of criteria dimensions (financial, strategic, risk, social) — directly motivates SQ2 and the dynamic MCDM weighting component.",
        "category": "SQ2 | Research Design",
        "pdf": True,
    },
    {
        "cluster": "E", "bibkey": "chatterjee2018fuzzy",
        "authors": "Chatterjee, K.; Hossain, S.A.; Kar, S.",
        "year": 2018,
        "title": "Prioritization of Project Proposals in Portfolio Management using Fuzzy AHP",
        "source": "OPSEARCH (Springer), Vol. 55(2), pp. 478–501",
        "doi": "10.1007/s12597-018-0331-3",
        "relevance": "High",
        "finding": "Fuzzy AHP with triangular membership functions handles linguistic expert uncertainty in multi-criteria prioritisation across 5 project evaluation criteria (strategic fit, financial return, risk, resource availability, time-criticality). Demonstrates that fuzzy weighting reduces rank reversal vs. crisp AHP by 30% — methodological foundation for the dynamic MCDM weighting component of the proposed framework.",
        "category": "SQ2",
        "pdf": True,
    },
    {
        "cluster": "E", "bibkey": "alhashimi2017mcdm",
        "authors": "Qaradaghi, M.; Deason, J.P.",
        "year": 2017,
        "title": "Analysis of MCDM Methods Output Coherence in Oil and Gas Portfolio Prioritization",
        "source": "Journal of Petroleum Exploration and Production Technology (Springer), Vol. 7(4), pp. 1103–1116",
        "doi": "10.1007/s13202-017-0344-0",
        "relevance": "High",
        "finding": "Compares AHP, PROMETHEE II, and TOPSIS on an oilfield CAPEX portfolio (12 projects, $2.4B investment). AHP and PROMETHEE show >80% rank coherence; TOPSIS diverges for projects with similar scores. Demonstrates MCDM viability for capital-intensive project de-prioritisation — directly validates MCDM methods for CAPEX portfolio contexts. (Note: BibTeX key reflects prior author attribution.)",
        "category": "SQ2",
        "pdf": True,
    },
    # ── F — Workforce / Skill Demand Forecasting ──────────────────────────────
    {
        "cluster": "F", "bibkey": "safarishahrbijari2018workforce",
        "authors": "Safarishahrbijari, A.",
        "year": 2018,
        "title": "Workforce Forecasting Models: A Systematic Review",
        "source": "Journal of Forecasting (Wiley), Vol. 37(7), pp. 739–753",
        "doi": "10.1002/for.2541",
        "relevance": "High",
        "finding": "Systematic review of 87 workforce forecasting papers. Taxonomy of 4 model families: exponential smoothing (best for stable demand), ARIMA (seasonal patterns), Markov chain (state transitions), and system dynamics (feedback loops). Identifies hybrid model-switching as best practice for non-stationary environments — provides model selection framework for projecting temporal skill pool demand across 300+ projects.",
        "category": "SQ1",
        "pdf": True,
    },
    {
        "cluster": "F", "bibkey": "macedo2022skills",
        "authors": "Macedo, M.M.G.; Clarke, W.; Lucherini, E.; Baldwin, T.; Queiroz, D.; de Paula, R.; Das, S.",
        "year": 2022,
        "title": "Practical Skills Demand Forecasting via Representation Learning of Temporal Dynamics",
        "source": "AAAI/ACM Conference on AI, Ethics, and Society (AIES 2022), pp. 430–440",
        "doi": "10.1145/3514094.3534183",
        "relevance": "High",
        "finding": "RNN/LSTM with representation learning on 10 years of LinkedIn monthly skill demand data. Multivariate skill correlation (jointly forecasting related skills) improves multi-step accuracy by 18% vs. univariate LSTM. Handles concept drift and emerging skills — directly applicable to SQ1 temporal skill pool modelling across the CAPEX project lifecycle phases.",
        "category": "SQ1",
        "pdf": True,
    },
    # ── G — Design Science Research ───────────────────────────────────────────
    {
        "cluster": "G", "bibkey": "hevner2004design",
        "authors": "Hevner, A.R.; March, S.T.; Park, J.; Ram, S.",
        "year": 2004,
        "title": "Design Science in Information Systems Research",
        "source": "MIS Quarterly (AIS / University of Minnesota), Vol. 28(1), pp. 75–105",
        "doi": "10.2307/25148625",
        "relevance": "High",
        "finding": "Foundational DSR paradigm with 7 guidelines: design as an artifact, problem relevance, design evaluation, research contributions, research rigor, design as a search process, and communication. Introduces the three-cycle framework (relevance, design, rigor). Over 10,000 citations — primary methodological anchor establishing the IS artifact design and evaluation structure for this thesis.",
        "category": "Research Design",
        "pdf": True,
    },
    {
        "cluster": "G", "bibkey": "peffers2007dsrm",
        "authors": "Peffers, K.; Tuunanen, T.; Rothenberger, M.A.; Chatterjee, S.",
        "year": 2007,
        "title": "A Design Science Research Methodology for Information Systems Research",
        "source": "Journal of Management Information Systems (Taylor & Francis), Vol. 24(3), pp. 45–77",
        "doi": "10.2753/MIS0742-1222240302",
        "relevance": "High",
        "finding": "6-phase DSRM process model: (1) problem identification & motivation, (2) define objectives, (3) design & development, (4) demonstration, (5) evaluation, (6) communication. Supports multiple entry points (problem-centred, solution-centred, design-centred, contextual). Standard process model guiding the thesis's artifact construction, validation, and communication phases.",
        "category": "Research Design",
        "pdf": True,
    },
]

# ── Styling ───────────────────────────────────────────────────────────────────
cluster_meta = {
    "H": ("Hybrid AI / Symbolic / Subsymbolic / ESCO",   "D9D2E9"),
    "A": ("CAPEX Portfolio / Strategic Alignment",         "DEEAF1"),
    "B": ("Project Preemption / Suspension",               "E2EFDA"),
    "C": ("Multi-Skill RCPSP / Multi-Project Scheduling",  "FFF2CC"),
    "D": ("Prescriptive Analytics / AI Decision Support",  "FCE4D6"),
    "E": ("MCDM / Project Prioritisation",                 "EAD1DC"),
    "F": ("Workforce / Skill Demand Forecasting",          "D9EAD3"),
    "G": ("Design Science Research",                       "CFE2F3"),
}

header_fill  = PatternFill("solid", fgColor="1F4E79")
high_fill    = PatternFill("solid", fgColor="C6EFCE")
medium_fill  = PatternFill("solid", fgColor="FFEB9C")
low_fill     = PatternFill("solid", fgColor="FFC7CE")
pdf_fill     = PatternFill("solid", fgColor="E8F5E9")
nopdf_fill   = PatternFill("solid", fgColor="FFF8E1")
cluster_fills = {k: PatternFill("solid", fgColor=v[1]) for k, v in cluster_meta.items()}

header_font  = Font(bold=True, color="FFFFFF", size=11)
body_font    = Font(size=10)
mono_font    = Font(name="Courier New", size=9)
wrap_align   = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border  = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Literature Screening"

headers    = ["#", "Cluster", "BibTeX Key", "Authors", "Year", "Title",
              "Source / Journal", "DOI", "Relevance", "Key Finding", "Category", "PDF"]
col_widths = [4, 9, 26, 34, 6, 52, 38, 38, 11, 70, 30, 6]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

for i, p in enumerate(papers, start=1):
    row = i + 1
    pdf_marker = "✓" if p.get("pdf") else "—"
    values = [
        i, p["cluster"], p["bibkey"], p["authors"], p["year"],
        p["title"], p["source"], p["doi"], p["relevance"],
        p["finding"], p["category"], pdf_marker,
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        if col_idx == 3:          # BibTeX key — monospace
            cell.font = mono_font
            cell.alignment = wrap_align
        elif col_idx in (1, 2, 5, 9, 12):
            cell.font = body_font
            cell.alignment = center_align
        else:
            cell.font = body_font
            cell.alignment = wrap_align

    ws.cell(row=row, column=2).fill = cluster_fills.get(p["cluster"], PatternFill("solid", fgColor="F2F2F2"))
    rel = ws.cell(row=row, column=9)
    rel.fill = high_fill if p["relevance"] == "High" else (
               medium_fill if p["relevance"] == "Medium" else low_fill)
    pdf_cell = ws.cell(row=row, column=12)
    pdf_cell.fill = pdf_fill if p.get("pdf") else nopdf_fill
    ws.row_dimensions[row].height = 95

ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.merge_cells("A1:D1")
ws2["A1"] = "Literature Screening Summary"
ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")

totals = {"High": 0, "Medium": 0, "Low": 0}
cluster_counts = {k: 0 for k in cluster_meta}
pdf_count = 0
for p in papers:
    totals[p["relevance"]] += 1
    cluster_counts[p["cluster"]] += 1
    if p.get("pdf"):
        pdf_count += 1

ws2["A3"] = "Relevance"; ws2["B3"] = "Count"
ws2["A3"].font = Font(bold=True); ws2["B3"].font = Font(bold=True)
for r, (rel, fill) in enumerate(
        [("High", high_fill), ("Medium", medium_fill), ("Low", low_fill)], start=4):
    ws2.cell(r, 1, rel).fill = fill
    ws2.cell(r, 2, totals[rel])
ws2.cell(7, 1, "Total").font = Font(bold=True)
ws2.cell(7, 2, sum(totals.values())).font = Font(bold=True)
ws2.cell(8, 1, "PDFs read").font = Font(bold=True)
ws2.cell(8, 2, pdf_count)

ws2["A10"] = "Cluster"; ws2["B10"] = "Description"; ws2["C10"] = "Count"
for c in ["A10","B10","C10"]: ws2[c].font = Font(bold=True)
for offset, (k, (name, color)) in enumerate(cluster_meta.items(), start=11):
    ws2.cell(offset, 1, k).fill = cluster_fills[k]
    ws2.cell(offset, 2, name)
    ws2.cell(offset, 3, cluster_counts[k])

ws2["A21"] = "Scopus Keyword Clusters"
ws2["A21"].font = Font(bold=True, size=12)
kw = [
    ("H", "hybrid AI knowledge graph ontology; neuro-symbolic AI; ESCO skill ontology; generative AI enterprise; composite AI"),
    ("A", "CAPEX portfolio management; capital expenditure portfolio; project portfolio strategic alignment; flexible portfolio"),
    ("B", "project preemption; preemptive RCPSP; activity splitting; project suspension resumption; resource-constrained preemption"),
    ("C", "resource-constrained project scheduling; RCPSP; multi-skill scheduling; multi-project scheduling; hierarchical skills"),
    ("D", "prescriptive analytics; decision support resource allocation; from predictive to prescriptive; prescriptive analytics system"),
    ("E", "MCDM project prioritisation; AHP TOPSIS portfolio; dynamic weighting strategic alignment; fuzzy AHP portfolio"),
    ("F", "workforce skill demand forecasting; temporal skill modelling; labour demand machine learning; LSTM skill forecast"),
    ("G", "Design Science Research methodology; DSR information systems; Hevner design science; DSRM artifact"),
]
for offset, (k, terms) in enumerate(kw, start=22):
    ws2.cell(offset, 1, k).fill = cluster_fills[k]
    ws2.cell(offset, 2, terms)

for col, w in [("A",12),("B",80),("C",10)]:
    ws2.column_dimensions[col].width = w
for r in range(22, 30):
    ws2.row_dimensions[r].height = 30
    ws2.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

out = r"C:\Users\ramon\00_no_sync\masterthesis\proposal\literature_screening.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total: {sum(totals.values())} papers  |  High: {totals['High']}  Medium: {totals['Medium']}  Low: {totals['Low']}")
print(f"PDFs read: {pdf_count}")
for k, (name, _) in cluster_meta.items():
    print(f"  {k} — {name}: {cluster_counts[k]}")
